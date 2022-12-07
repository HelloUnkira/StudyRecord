# 百度翻译脚本:自动化流程+同步
from urllib.request import Request
from urllib.request import urlopen
from urllib.parse import urlencode
import os
import json
import execjs
import openpyxl


if __name__ == '__main__':
    print('\n\n-----start-----\n\n')
    # 1.加载百度翻译JS脚本
    print('\n\n-----step1-----\n\n')
    with open('reverse.js', mode='r', encoding='utf-8') as js_file:
        js_cmd = execjs.compile(js_file.read())
    # 2.加载本地浏览器请求信息并生成请求头,以及数据包定向字段
    print('\n\n-----step2-----\n\n')
    with open('translate.json', mode='r', encoding='utf-8') as json_file:
        json_data = json.loads(json_file.read())
        print('User-Agent: ', json_data['User-Agent'])
        print('Cookie:     ', json_data['Cookie'])
        print('Acs-Token:  ', json_data['Acs-Token'])
        print('token:      ', json_data['token'])
    headers = {
        'User-Agent': json_data['User-Agent'],
        'Cookie': json_data['Cookie'],
        'Referer': 'https://fanyi.baidu.com',
        'Acs-Token': json_data['Acs-Token'],
        'x-requested-with': 'XMLHttpRequest'}
    token = json_data['token']
    # 3.读取输入类型, 输入源, 输出类型列表
    print('\n\n-----step3-----\n\n')
    input_type = json_data['input_type']
    with open(json_data['input_file'], mode='r', encoding='utf-8') as input_file:
        input_list = input_file.read().split(',')
    output_list = json_data['output_list']
    print('input_type:   ', input_type)
    print('input_list:   ', input_list)
    print('output_list:  ', output_list)
    print('output_types: ', len(output_list))
    # 4.启用翻译器开始翻译,生成翻译列表集合
    print('\n\n-----step4-----\n\n')
    translate_list = []
    process_current = 1
    process_target = len(input_list) * len(output_list)
    # 在翻译时,如果捕获到了相同的类型的目标,直接跳过
    for i, word in enumerate(input_list):
        word_list = []
        for j, output_type in enumerate(output_list):
            print('process:---%d/%d---' %(process_current, process_target))
            if input_type == output_type:
                result = word
            else:
                # 计算传输密钥
                sign = js_cmd.call('tl', word)
                # 第一次请求通讯
                url = 'https://fanyi.baidu.com/langdetect'
                data = {'query': word}
                with urlopen(Request(url, data=urlencode(data).encode('utf-8'), headers=headers)) as response:
                    assert response.code == 200
                    response_data = json.loads(response.read().decode('utf-8'))
                    if response_data['msg'] != 'success':
                        print(response_data, '\t', sign)
                # 第二次请求通讯
                url = 'https://fanyi.baidu.com/v2transapi?'
                data = {'from': input_type, 'to': output_type, 'query': word, 'transtype': 'realtime',
                        'simple_means_flag': '3', 'sign': sign,
                        'token': token, 'domain': 'common'}
                with urlopen(Request(url, data=urlencode(data).encode('utf-8'), headers=headers)) as response:
                    assert response.code == 200
                    response_data = json.loads(response.read().decode('utf-8'))
                    try:
                        result = response_data['trans_result']['data'][0]['dst']
                    except Exception as E:
                        print('Catch Error:', response_data)
                        result = ''
            process_current += 1
            word_list.append(result)
        translate_list.append(word_list)
    # 5.将数据按照格式载入到磁盘中持久化
    print('\n\n-----step5-----\n\n')
    # 将其按列表的每一个行持久化到一个文本文件中,
    with open('translate.output', mode='w', encoding='utf-8') as output_file:
        for word_list in translate_list:
            output_file.write(str(word_list) + '\n')
    if os.path.isfile('translate.xlsx'):
        os.remove('translate.xlsx')
    output_file = openpyxl.Workbook()
    output_file.create_sheet('translate', 0)
    for i, line in enumerate(translate_list):
        for j, word in enumerate(translate_list[i]):
            output_file['translate'].cell(i + 1, j + 1).value = word
    output_file.save('translate.xlsx')
    print('\n\n-----end-----\n\n')

