# 百度翻译脚本:自动化流程+异步协程
import asyncio
import aiohttp
import os
import json
import openpyxl
import time
import hashlib


def list_split_block(iterator, unit=1):
    length = len(iterator)
    for index in range(0, length, unit):
        yield iterator[index: min(index + unit, length)]


async def asyncio_translate(data_set) -> None:
    transfer_session = data_set['session']
    language_from = data_set['input_type']
    language_to = data_set['output_type']
    data_query = data_set['word']
    data_list = data_set['list']
    data_index1 = data_set['index_1']
    data_index2 = data_set['index_2']
    record_process = data_set['translate_process']
    # 草，有空去爆破它
    # 百度API申请信息, 密钥
    baiduAPI_ID  = data_set['api_id']
    baiduAPI_Key = data_set['api_key']
    url = "http://api.fanyi.baidu.com/api/trans/vip/translate"
    salt = str(time.time())
    sign = hashlib.md5((baiduAPI_ID + data_query + salt + baiduAPI_Key).encode('utf-8')).hexdigest()
    transfer_data = {'q': data_query, 'from': language_from, 'to': language_to,
                     'appid': baiduAPI_ID, 'salt': salt, 'sign': sign}
    async with transfer_session.post(url, data=transfer_data) as response:
        assert response.status == 200
        response.encoding = 'utf-8'
        response_data = json.loads(await response.content.read())
        try:
            translate_result = response_data['trans_result'][0]['dst']
        except Exception as error:
            print('catch error:', error, response_data)
            translate_result = ''

    data_list[data_index1][data_index2] = translate_result
    record_process[0] += 1
    print('process:---%d/%d---' % (record_process[0], record_process[1]))


async def main_process() -> None:
    print('\n\n-----start-----\n\n')
    # 1.加载百度翻译JS脚本
    print('\n\n-----step1-----\n\n')
    # 2.加载本地浏览器请求信息并生成请求头,以及数据包定向字段
    print('\n\n-----step2-----\n\n')
    with open('translate.json', mode='r', encoding='utf-8') as json_file:
        json_data = json.loads(json_file.read())
    # 3.读取输入类型, 输入源, 输出类型列表
    print('\n\n-----step3-----\n\n')
    input_type = json_data['input_type']
    with open(json_data['input_file'], mode='r', encoding='utf-8') as input_file:
        input_list = [word for word in input_file.read().split(',') if word != '']
    output_list = json_data['output_list']
    print('input_type:   ', input_type)
    print('input_list:   ', input_list)
    print('output_list:  ', output_list)
    print('output_types: ', len(output_list))
    # 4.启用翻译器开始翻译,生成翻译列表集合
    print('\n\n-----step4-----\n\n')
    translate_list = [['' for j in range(0, len(output_list))] for i in range(0, len(input_list))]
    translate_process = [0, len(input_list) * len(output_list)]
    # 在翻译时,如果捕获到了相同的类型的目标,直接跳过
    for i, word in enumerate(input_list):
        for j, output_type in enumerate(output_list):
            if input_type == output_type:
                translate_list[i][j] = word
                translate_process[0] += 1
    # 在翻译时,如果捕获到了不相同的类型的目标,生成执行队列
    async with aiohttp.ClientSession() as session:
        # 生成所有的工作列表,但是并不去执行它,而是在下方使用同步分组的方式去执行
        translate_work_list = [asyncio_translate({'input_type': input_type, 'output_type': output_type,
                                                  'session': session, 'word': word,
                                                  'list': translate_list, 'index_1': i, 'index_2': j,
                                                  'translate_process': translate_process,
                                                  'api_id': json_data['api_id'], 'api_key': json_data['api_key']})
                               for i, word in enumerate(input_list)
                               for j, output_type in enumerate(output_list)
                               if input_type != output_type]
        print('---async task ready---')
        # 为了解决异步协程请求密集频繁的问题导致服务器拒绝响应,这里需要手动对其进行一个降速
        # 我们将其按照一定的比例将任务切分成任务块,一次单元内仅仅允许指定的任务去运行
        for work_unit_list in list_split_block(translate_work_list, int(json_data['await_unit'])):
            print('---one unit await---')
            await asyncio.wait([asyncio.create_task(work) for work in work_unit_list])
            print('---one unit over---')
            await asyncio.sleep(1)

    # 5.将数据按照格式载入到磁盘中持久化
    print('\n\n-----step5-----\n\n')
    # 将其按列表的每一个行持久化到一个文本文件中,
    with open('translate.output', mode='w', encoding='utf-8') as output_file:
        for word_list in translate_list:
            output_file.write(str(word_list) + '\n')
    if os.path.isfile('translate.xlsx'):
        os.remove('translate.xlsx')
    output_file = openpyxl.Workbook()
    output_file.create_sheet('translate', 0)
    for i, line in enumerate(translate_list):
        for j, word in enumerate(translate_list[i]):
            output_file['translate'].cell(i + 1, j + 1).value = word
    output_file.save('translate.xlsx')
    print('\n\n-----end-----\n\n')


if __name__ == '__main__':
    translate_task_loop = asyncio.get_event_loop()
    translate_task_loop.run_until_complete(main_process())

